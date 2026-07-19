"""Assemble parsed IEX jsonl into an Excel workbook for personal records.

Two sheets (DAM, RTM), one row per 15-min block, sorted by date then block.
DAM has ~490k rows and RTM ~210k — both under Excel's 1,048,576-row/sheet cap,
so each fits on a single sheet.
"""
import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

SHEETS = {
    "DAM (15-min blocks)": ("dam", [
        "report_date", "block", "hour", "time_block", "purchase_bid_mw",
        "sell_bid_mw", "mcv_mw", "final_scheduled_volume_mw", "mcp_rs_mwh"]),
    "RTM (15-min blocks)": ("rtm", [
        "report_date", "block", "hour", "session_id", "time_block",
        "purchase_bid_mw", "sell_bid_mw", "mcv_mw", "final_scheduled_volume_mw",
        "mcp_rs_mwh"]),
}


def load(key):
    rows = [json.loads(l) for l in open(f"iex_{key}.jsonl")]
    rows.sort(key=lambda r: (r["report_date"], r["block"]))
    return rows


wb = Workbook()
wb.remove(wb.active)
for sheet_name, (key, cols) in SHEETS.items():
    ws = wb.create_sheet(title=sheet_name)
    ws.append([c.replace("_", " ").title() for c in cols])
    ws.freeze_panes = "A2"
    for r in load(key):
        ws.append([r.get(c) for c in cols])
    for i, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(11, min(26, len(c) + 2))
    print(f"{sheet_name}: {ws.max_row - 1} rows")

out = "../../IEX_Historical_DAM_RTM_2012-2026.xlsx"
wb.save(out)
print("saved", out)
