"""Assemble the parsed NLDC jsonl (main + recovered) into one multi-sheet
xlsx workbook for personal records. One sheet per table, sorted by date."""
import json
import glob
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

REGION_RANK = {"NR": 0, "WR": 1, "SR": 2, "ER": 3, "NER": 4, "TOTAL": 5, "ALL_INDIA": 5}

SHEETS = {
    "Regional PSP (A)": ("regional", [
        "report_date", "region", "demand_met_evening_peak_mw", "peak_shortage_mw",
        "energy_met_mu", "hydro_gen_mu", "wind_gen_mu", "solar_gen_mu",
        "energy_shortage_mu", "max_demand_met_mw", "max_demand_time", "source_format"]),
    "State PSP (C)": ("state", [
        "report_date", "region", "state_raw", "state_canonical", "max_demand_met_mw",
        "shortage_during_max_demand_mw", "energy_met_mu", "drawal_schedule_mu",
        "od_ud_mu", "max_od_import_mw", "max_od_export_mw", "energy_shortage_mu",
        "source_format"]),
    "Generation Outage (F)": ("gen_outage", [
        "report_date", "region", "central_sector_mw", "state_sector_mw", "total_mw",
        "source_format"]),
    "Sourcewise Generation (G)": ("sourcewise_gen", [
        "report_date", "region", "coal_mu", "lignite_mu", "thermal_combined_mu",
        "hydro_mu", "nuclear_mu", "gas_mu", "res_mu", "total_mu", "res_share_pct",
        "non_fossil_share_pct", "source_format"]),
    "Solar vs Non-Solar Peak (I)": ("solar_nonsolar", [
        "report_date", "hour_type", "max_demand_met_mw", "max_demand_time",
        "shortage_mw", "source_format"]),
}


def load(key):
    rows = []
    for fn in [f"nldc_{key}.jsonl", f"nldc_{key}_recovered.jsonl"]:
        try:
            for line in open(fn):
                rows.append(json.loads(line))
        except FileNotFoundError:
            pass
    rows.sort(key=lambda r: (r["report_date"],
                             REGION_RANK.get(r.get("region", ""), 9),
                             r.get("state_canonical", ""),
                             r.get("hour_type", "")))
    return rows


wb = Workbook()
wb.remove(wb.active)
for sheet_name, (key, cols) in SHEETS.items():
    ws = wb.create_sheet(title=sheet_name)
    ws.append([c.replace("_", " ").title() for c in cols])
    ws.freeze_panes = "A2"
    for r in load(key):
        ws.append([r.get(c) for c in cols])
    # reasonable column widths
    for i, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, min(26, len(c) + 2))
    print(f"{sheet_name}: {ws.max_row - 1} rows")

out = "../../NLDC_Historical_PSP_2013-2026.xlsx"
wb.save(out)
print("saved", out)
