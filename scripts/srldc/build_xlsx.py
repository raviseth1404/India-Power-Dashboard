"""Assemble the parsed SRLDC jsonl into one multi-sheet xlsx workbook for
personal records. One sheet per table, sorted by date."""
import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

SHEETS = {
    "Regional Availability (T1)": ("regional", [
        "report_date", "evening_peak_demand_met_mw", "evening_peak_shortage_mw",
        "evening_peak_requirement_mw", "evening_peak_freq_hz",
        "offpeak_demand_met_mw", "offpeak_shortage_mw", "offpeak_requirement_mw",
        "offpeak_freq_hz", "day_energy_demand_met_mu", "day_energy_shortage_mu"]),
    "State Demand Met (T2C)": ("state_demand", [
        "report_date", "state_canonical", "state_raw", "max_demand_met_mw",
        "max_demand_time", "shortage_at_max_demand_mw", "requirement_at_max_demand_mw",
        "demand_met_at_max_req_mw", "max_req_time", "shortage_at_max_req_mw",
        "max_requirement_mw", "ace_max", "ace_max_time", "ace_min", "ace_min_time"]),
    "State Generation (T3A)": ("state_generation", [
        "report_date", "state_canonical", "seq", "row_type", "source_category",
        "station_name", "inst_capacity_mw", "peak_2000_mw", "offpeak_0300_mw",
        "daypeak_mw", "daypeak_time", "mingen_mw", "mingen_time", "gross_gen_mu",
        "net_gen_mu", "avg_mw_0618"]),
}


def load(key):
    rows = [json.loads(l) for l in open(f"srldc_{key}.jsonl")]
    rows.sort(key=lambda r: (r["report_date"], r.get("state_canonical", ""),
                             r.get("seq", 0)))
    return rows


wb = Workbook()
wb.remove(wb.active)
for sheet_name, (key, cols) in SHEETS.items():
    ws = wb.create_sheet(title=sheet_name)
    ws.append([c.replace("_", " ").title() for c in cols])
    ws.freeze_panes = "A2"
    for r in load(key):
        ws.append([r.get(c) for c in cols])
    for i, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, min(28, len(c) + 2))
    print(f"{sheet_name}: {ws.max_row - 1} rows")

out = "../../SRLDC_Historical_PSP_2018-2026.xlsx"
wb.save(out)
print("saved", out)
